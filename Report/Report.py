"""
定义write_report_ETH(file_path,file_name,date,qr,strategy) 和 write_report_XBT(file_path,file_name,date,qr,strategy)
通过从数据库中查询数据，新建报表并写入数据
配合模板使用和维护，模板文件：daily_silp_point_analysis_ETH_template.xlsx 和 daily_silp_point_analysis_XBT_template.xlsx
luolan
"""

from shutil import copyfile
from DataQuery import query
import openpyxl
import datetime as dt



def write_report_ETH(file_path,file_name,date,qr,strategy):
    """
    :param file_path: 生成报表后的存放路径
    :param file_name: 报表文件名
    :param date: 报表日期
    :param qr: DataQuery.query.query()
    :param strategy: 策略名称，例如：bitmex_4a_ETH_Balance
    """

    copyfile("daily_silp_point_analysis_ETH_template.xlsx",file_path+file_name)
    wb = openpyxl.load_workbook(file_path+file_name)
    ws = wb['data']

    ws['B3'] = date+" 00:00:00"
    ws['B4'] = date+" 23:59:59"
    datetime = date+" 00:00:00"
    ws['B5'] = qr.get_price(datetime,'eth','open')
    ws['B9'] = qr.get_price(datetime,'xbt','open')
    datetime = date+" 23:59:00"
    ws['B6'] = qr.get_price(datetime,'eth','close')
    ws['B10'] = qr.get_price(datetime,'xbt','close')
    ws['B7'] = strategy
    datetime = date+ " 00:00:00"
    ws['B8'] = qr.get_position(datetime,strategy)
    begin_datetime = date+" 00:00:00"
    end_datetime = (dt.datetime.strptime(begin_datetime,"%Y-%m-%d %H:%M:%S")+dt.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
    trade_event = qr.get_trade_event(begin_datetime, end_datetime, strategy)
    if trade_event is not None:
        #写入交易记录
        for i in range(len(trade_event)):
            for j in range(7):
                ws.cell(row = i+12, column = 1+j).value = trade_event[i][j]
        #写入对应时间xbt_price 和上一分钟的 eth_price
        for i in range(len(trade_event)):
            datetime = trade_event[i][5]
            # datetime = dt.datetime.strptime(datetime,"%Y-%m-%d %H:%M:%S")
            datetime_now = datetime + dt.timedelta(seconds=-datetime.second)
            datetime_last_m = datetime_now + dt.timedelta(minutes=-1)
            datetime_now = datetime_now.strftime("%Y-%m-%d %H:%M:%S")
            datetime_last_m = datetime_last_m.strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row = i+12, column = 8).value = qr.get_price(datetime_now,'xbt','open')
            ws.cell(row = i+12, column = 9).value = qr.get_price(datetime_last_m,'eth','close')

    wb.save(file_path + file_name)


def write_report_XBT(file_path,file_name,date,qr,strategy):
    """
    :param file_path: 生成报表后的存放路径
    :param file_name: 报表文件名
    :param date: 报表日期
    :param qr: DataQuery.query.query()
    :param strategy: 策略名称，例如：bitmex_4a_BTC_Balance
    """

    copyfile("daily_silp_point_analysis_XBT_template.xlsx",file_path+file_name)
    wb = openpyxl.load_workbook(file_path+file_name)
    ws = wb['data']

    ws['B3'] = date+" 00:00:00"
    ws['B4'] = date+" 23:59:59"
    datetime = date+" 00:00:00"
    ws['B5'] = qr.get_price(datetime,'xbt','open')
    datetime = date+" 23:59:00"
    ws['B6'] = qr.get_price(datetime,'xbt','close')
    ws['B7'] = strategy
    datetime = date+ " 00:00:00"
    ws['B8'] = qr.get_position(datetime,strategy)
    begin_datetime = date+" 00:00:00"
    end_datetime = (dt.datetime.strptime(begin_datetime,"%Y-%m-%d %H:%M:%S")+dt.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
    trade_event = qr.get_trade_event(begin_datetime, end_datetime, strategy)
    if trade_event is not None:
        #写入交易记录
        for i in range(len(trade_event)):
            for j in range(7):
                ws.cell(row = i+10, column = 1+j).value = trade_event[i][j]
        #写入上一分钟的 xbt_price
        for i in range(len(trade_event)):
            datetime = trade_event[i][5]
            datetime_now = datetime + dt.timedelta(seconds=-datetime.second)
            datetime_last_m = datetime_now + dt.timedelta(minutes=-1)
            datetime_last_m = datetime_last_m.strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row = i+10, column = 8).value = qr.get_price(datetime_last_m,'xbt','close')

    wb.save(file_path + file_name)




