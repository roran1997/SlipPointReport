'''
2、3、4、5月份 旧模板报表重做脚本
luolan
'''

import datetime as dt
from DataQuery import query
import openpyxl
from shutil import copyfile

def copy_trade_event_ETH(source_file:openpyxl.workbook, target_file:openpyxl.workbook, maxrows=200):
    ws_s = source_file['data']
    ws_t = target_file['data']
    for i in range(maxrows):
        for j in range(8):
            ws_t.cell(row=12+i,column=j+1).value = ws_s.cell(row=12+i,column=j+1).value
    for i in range(8):
        ws_t.cell(row=3+i,column=2).value = ws_s.cell(row=3+i,column=2).value


def copy_trade_event_XBT(source_file, target_file, maxrows=200):
    ws_s = source_file['data']
    ws_t = target_file['data']
    for i in range(maxrows):
        for j in range(7):
            ws_t.cell(row=10 + i, column=j + 1).value = ws_s.cell(row=10 + i, column=j + 1).value
    for i in range(6):
        ws_t.cell(row=3 + i, column=2).value = ws_s.cell(row=3 + i, column=2).value


def find_net_value_row(ws,col):
    cur_row = 1
    while(ws.cell(row=cur_row,column=col).value != "净资产"):
        cur_row  = cur_row+1
    # print(ws.cell(row=cur_row,column=col).value)
    return cur_row

def write_ETH_report(source_file,target_file,qr:query):
    copyfile("daily_silp_point_analysis_ETH_template.xlsx", target_file)
    wb_s = openpyxl.load_workbook(source_file,data_only=True)
    wb_t = openpyxl.load_workbook(target_file)
    copy_trade_event_ETH(wb_s,wb_t)
    ws_t = wb_t['data']
    cur_row = 12
    while(1):
        if(ws_t.cell(row=cur_row, column=1).value is None): break
        datetime = ws_t.cell(row=cur_row,column=6).value
        if isinstance(datetime,str):
            datetime = dt.datetime.strptime(datetime,"%Y-%m-%d %H:%M:%S")
        datetime_now = datetime + dt.timedelta(seconds=-datetime.second)
        datetime_last_m = datetime_now + dt.timedelta(minutes=-1)
        datetime_last_m = datetime_last_m.strftime("%Y-%m-%d %H:%M:%S")
        ws_t.cell(row=cur_row,column=9).value = qr.get_price(datetime_last_m,'eth','close')
        cur_row +=1
    ws_t = wb_t['Balance_stats']
    ws_s = wb_s['Balance_stats']
    net_value_row = find_net_value_row(ws_s,4)
    # print(net_value_row)
    ws_t['E510'] = ws_s.cell(row=net_value_row-1,column=5).value
    ws_t['E507'] = ws_s.cell(row=net_value_row-4,column=5).value

    wb_s.close()
    wb_t.save(target_file)

def write_XBT_report(source_file,target_file,qr:query):
    copyfile("daily_silp_point_analysis_XBT_template.xlsx", target_file)
    wb_s = openpyxl.load_workbook(source_file,data_only=True)
    wb_t = openpyxl.load_workbook(target_file)
    copy_trade_event_XBT(wb_s,wb_t)
    ws_t = wb_t['data']
    cur_row = 10
    while(1):
        if ws_t.cell(row=cur_row,column=1).value is None: break
        datetime = ws_t.cell(row=cur_row,column=6).value
        if isinstance(datetime,str):
            datetime = dt.datetime.strptime(datetime,"%Y-%m-%d %H:%M:%S")
        datetime_now = datetime + dt.timedelta(seconds=-datetime.second)
        datetime_last_m = datetime_now + dt.timedelta(minutes=-1)
        datetime_last_m = datetime_last_m.strftime("%Y-%m-%d %H:%M:%S")
        ws_t.cell(row=cur_row,column=8).value = qr.get_price(datetime_last_m,'xbt','close')
        cur_row +=1
    ws_t = wb_t['Balance_stats']
    ws_s = wb_s['Balance_stats']
    net_value_row = find_net_value_row(ws_s,4)
    # print(net_value_row)
    ws_t['E510'] = ws_s.cell(row=net_value_row-1,column=5).value
    ws_t['E507'] = ws_s.cell(row=net_value_row-4,column=5).value

    wb_s.close()
    wb_t.save(target_file)



f_name = "note2.txt"
qr = query.query(f_name=f_name)

# #bitmex_4a_2月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4a/2019年2月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4a/2019年2月/"
# date = dt.date(2019,2,20)
# date_end = dt.date(2019,2,28)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")

    # #ETH
    # target_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
    # source_file_name = "daily_silp_point_analysis_ETH_"+date1+".xlsx"
    # try:
    #     write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
    # except:
    #     f = open(f_name,'a')
    #     f.write("error: "+ target_file_name + '\n')
    #     f.close()

    # #XBT
    # target_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
    # source_file_name = "daily_silp_point_analysis_XBT_" + date1 + ".xlsx"
    #
    # write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
    # date = date + dt.timedelta(days=1)

#
# #bitmex_4a_3月 part1
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4a/2019年3月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4a/2019年3月/"
# date = dt.date(2019,3,1)
# date_end = dt.date(2019,3,10)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "daily_silp_point_analysis_ETH_"+date1+".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "daily_silp_point_analysis_XBT_" + date1 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#     date = date + dt.timedelta(days=1)


# #bitmex_4a_3月 part 2
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4a/2019年3月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4a/2019年3月/"
# date = dt.date(2019,3,11)
# date_end = dt.date(2019,3,31)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     # try:
#     #     write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     # except:
#     #     f = open(f_name,'a')
#     #     f.write("error: "+ target_file_name + '\n')
#     #     f.close()
#     write_ETH_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#
#
#     #XBT
#     target_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     # try:
#     #     write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     # except:
#     #     f = open(f_name, 'a')
#     #     f.write("error: " + target_file_name + '\n')
#     #     f.close()
#     write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#
#     date = date + dt.timedelta(days=1)
#
# #bitmex_4a_4月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4a/2019年4月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4a/2019年4月/"
# date = dt.date(2019,4,1)
# date_end = dt.date(2019,4,30)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#     date = date + dt.timedelta(days=1)
# #bitmex_4a_5月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4a/2019年5月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4a/2019年5月/"
# date = dt.date(2019,5,1)
# date_end = dt.date(2019,5,28)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#
#     date = date + dt.timedelta(days=1)


#

# #bitmex_4b_3月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4b/2019年3月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4b/2019年3月/"
# date = dt.date(2019,3,14)
# date_end = dt.date(2019,3,31)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_4b_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4b_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_4b_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4b_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#
#     date = date + dt.timedelta(days=1)

# #bitmex_4b_4月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4b/2019年4月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4b/2019年4月/"
# date = dt.date(2019,4,1)
# date_end = dt.date(2019,4,30)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_4b_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4b_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_4b_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4b_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#
#     date = date + dt.timedelta(days=1)

# #bitmex_4b_5月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_4b/2019年5月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_4b/2019年5月/"
# date = dt.date(2019,5,1)
# date_end = dt.date(2019,5,28)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_4b_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4b_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_4b_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_4b_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#
#     date = date + dt.timedelta(days=1)
#
# #bitmex_666a_3月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_666a/2019年3月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_666a/2019年3月/"
# date = dt.date(2019,3,14)
# date_end = dt.date(2019,3,31)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_666a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_666a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_666a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_666a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#
#     date = date + dt.timedelta(days=1)
#
# #bitmex_666a_4月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_666a/2019年4月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_666a/2019年4月/"
# date = dt.date(2019,4,1)
# date_end = dt.date(2019,4,30)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_666a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_666a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_666a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_666a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#
#     date = date + dt.timedelta(days=1)
#
# #bitmex_666a_5月
# target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_666a/2019年5月/"
# source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_666a/2019年5月/"
# date = dt.date(2019,5,1)
# date_end = dt.date(2019,5,28)
# while(date<=date_end):
#     date1 = date.strftime("%Y%m%d")
#     date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
#
#     #ETH
#     target_file_name = "bitmex_666a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_666a_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
#     except:
#         f = open(f_name,'a')
#         f.write("error: "+ target_file_name + '\n')
#         f.close()
#
#     #XBT
#     target_file_name = "bitmex_666a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     source_file_name = "bitmex_666a_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
#     try:
#         write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
#     except:
#         f = open(f_name, 'a')
#         f.write("error: " + target_file_name + '\n')
#         f.close()
#
#     date = date + dt.timedelta(days=1)
#

#bitmex_5_5月
target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_5/2019年5月/"
source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_5/2019年5月/"
date = dt.date(2019,5,20)
date_end = dt.date(2019,5,28)
while(date<=date_end):
    date1 = date.strftime("%Y%m%d")
    date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")

    #ETH
    target_file_name = "bitmex_5_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
    source_file_name = "bitmex_5_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
    try:
        write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
    except:
        f = open(f_name,'a')
        f.write("error: "+ target_file_name + '\n')
        f.close()

    #XBT
    target_file_name = "bitmex_5_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
    source_file_name = "bitmex_5_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
    try:
        write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
    except:
        f = open(f_name, 'a')
        f.write("error: " + target_file_name + '\n')
        f.close()

    date = date + dt.timedelta(days=1)


#bitmex_a1_5月
target_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/bitmex_a1/2019年5月/"
source_file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(old)/bitmex_a1/2019年5月/"
date = dt.date(2019,5,21)
date_end = dt.date(2019,5,28)
while(date<=date_end):
    date1 = date.strftime("%Y%m%d")
    date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")

    #ETH
    target_file_name = "bitmex_a1_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
    source_file_name = "bitmex_a1_ETH_Balance_" + date1 + "_" + date2 + ".xlsx"
    try:
        write_ETH_report(source_file_path+source_file_name,target_file_path+target_file_name,qr)
    except:
        f = open(f_name,'a')
        f.write("error: "+ target_file_name + '\n')
        f.close()

    #XBT
    target_file_name = "bitmex_a1_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
    source_file_name = "bitmex_a1_BTC_Balance_" + date1 + "_" + date2 + ".xlsx"
    try:
        write_XBT_report(source_file_path + source_file_name, target_file_path + target_file_name, qr)
    except:
        f = open(f_name, 'a')
        f.write("error: " + target_file_name + '\n')
        f.close()

    date = date + dt.timedelta(days=1)

