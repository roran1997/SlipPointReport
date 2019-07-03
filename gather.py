'''
汇总每月的滑点统计
luolan
openpyxl.load_workbook（file name) 中不可以加载含图的workbook，所以先暂时将带图的sheet移动到空白工作簿，程序结束后再移动回来
'''

import datetime as dt
import openpyxl

def find_row(ws, s:str, col:int):
    i = 1
    while(i<=1048756):
        if ws.cell(row=i,column=col).value == s:
            return i
        i += 1
    return None

def copy(ws_s,ws_t,row_s:int,row_t,columns:int):
    for i in range(columns):
        ws_t.cell(row=row_t,column=3+i).value = ws_s.cell(row=row_s,column=2+i).value

account_list = ["bitmex_4a","bitmex_4b","bitmex_666a","bitmex_a1","bitmex_5","bitmex_5a"]
# account_list = ["bitmex_4a","bitmex_4b","bitmex_666a"]
#变量：
file_path = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/"
file_path_s_month = "/2019年6月/"
file_name = "6月汇总.xlsx"
date_start = dt.date(2019,6,1)
date_end = dt.date(2019, 6, 15)


wb = openpyxl.load_workbook(file_path+file_name)
s = "Total"

for account in account_list:
    file_path_s = "C:/Users/lmy/Desktop/Bitmex/daily_slip_point_analysis(new)/" + account + file_path_s_month
    date = date_start
    #BTC
    strategy = account+"_BTC"
    ws = wb[strategy]
    cur_row = 2
    while (date <= date_end):
        date1 = date.strftime("%Y%m%d")
        date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
        file_name_s = strategy+"_Balance_"+date1+"_"+date2+".xlsx"
        try:
            wb_s = openpyxl.load_workbook(file_path_s+file_name_s,data_only=True)
            ws_s = wb_s["Balance_stats"]
            target_row = find_row(ws_s,s=s,col=1)
            copy(ws_s,ws,target_row,cur_row,18)
            wb_s.close()
        except:
            print(strategy+file_name_s)
        ws.cell(row=cur_row,column=1).value = strategy
        ws.cell(row=cur_row, column=2).value = date.strftime("%Y-%m-%d")
        cur_row +=1
        date = date + dt.timedelta(days=1)

    # ETH
    date = date_start
    strategy = account + "_ETH"
    ws = wb[strategy]
    cur_row = 2
    while (date <= date_end):
        date1 = date.strftime("%Y%m%d")
        date2 = (date + dt.timedelta(days=1)).strftime("%Y%m%d")
        file_name_s = strategy + "_Balance_" + date1 + "_" + date2 + ".xlsx"
        try:
            wb_s = openpyxl.load_workbook(file_path_s + file_name_s, data_only=True)
            ws_s = wb_s["Balance_stats"]
            target_row = find_row(ws_s, s=s, col=1)
            copy(ws_s, ws, target_row, cur_row, 18)
            wb_s.close()
        except:
            print(strategy+file_name_s)
        ws.cell(row=cur_row, column=1).value = strategy
        ws.cell(row=cur_row, column=2).value = date.strftime("%Y-%m-%d")
        cur_row += 1
        date = date + dt.timedelta(days=1)
    wb.save(file_path+file_name)







